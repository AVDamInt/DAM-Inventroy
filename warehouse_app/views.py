import json

import pandas as pd
from django.contrib.auth.decorators import login_required
from django.core import serializers
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.views import generic
from django.views.generic import ListView
from django.db.models import Q

from . import models
from django.urls import reverse_lazy
from django.views.generic.edit import DeleteView, CreateView
from . import forms
from . import utils
from . import filters
from django_filters.views import FilterView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.edit import UpdateView
import xlwt
from openpyxl import Workbook
from django.contrib import messages


class SearchResultsListView(ListView):  # new
    model = models.Device
    context_object_name = "device_list"
    template_name = "search_results.html"

    def get_queryset(self):  # new
        query = self.request.GET.get("q")
        return models.Device.objects.filter(
            Q(serial_number__icontains=query) | Q(contract__icontains=query) | Q(host_name__icontains=query)
        )

class DepartmentList(LoginRequiredMixin, FilterView):
    # login_url = 'accounts/login'
    # redirect_field_name = 'redirect_to'
    paginate_by = 10
    context_object_name = "departments"
    model = models.Department
    template_name = "department_list.html"


@login_required(login_url="/accounts/login/")
def create_department(request):
    if request.method == "POST":
        instance_form = forms.DepartmentForm(request.POST)
        if instance_form.is_valid():
            instance_form.save()
            return HttpResponseRedirect(reverse_lazy("department_list"))
    else:
        form = forms.DepartmentForm()
        return render(request, "department_form.html", {"form": form})


class DepartmentDetailView(LoginRequiredMixin, generic.DetailView):
    login_url = "accounts/login"
    model = models.Department
    template_name = "department_detail.html"
    context_object_name = "department"


class DepartmentUpdateView(LoginRequiredMixin, generic.UpdateView):
    template_name = "department_update_form.html"
    model = models.Department
    form_class = forms.DepartmentUpdateForm


class DepartmentDeleteView(LoginRequiredMixin, DeleteView):
    # login_url = 'accounts/login'
    model = models.Department
    template_name = "department_delete.html"
    success_url = reverse_lazy("department_list")



@login_required(login_url="/accounts/login/")
def create_device(request):
    if request.method == "POST":
        instance_form = forms.DeviceForm(request.POST)
        if instance_form.is_valid():
            instance_form.save()
            return HttpResponseRedirect(reverse_lazy("device_list"))
    else:
        form = forms.DeviceForm()
        return render(request, "device_form.html", {"form": form})




class DeviceList(LoginRequiredMixin, FilterView):
    # login_url = 'accounts/login'
    # redirect_field_name = 'redirect_to'
    paginate_by = 10
    context_object_name = "devices"
    model = models.Device
    template_name = "device_list.html"
    filterset_class = filters.DeviceFilter

    def get_context_data(self, **kwargs):
        context = super(DeviceList, self).get_context_data(**kwargs)
        context["hist_devices"] = (models.Device.objects.filter(history_type=0)).count()
        context["active_devices"] = (
            models.Device.objects.filter(history_type=1)
        ).count()
        context["available_devices"] = (models.Device.objects.filter(status=0)).count()
        context["unavailable_devices"] = (
            models.Device.objects.filter(status=1)
        ).count()
        return context


class DeviceDetailView(LoginRequiredMixin, generic.DetailView):
    # login_url = 'accounts/login'
    model = models.Device
    template_name = "device_detail.html"
    context_object_name = "device"

    def get_context_data(self, **kwargs):
        context = super(DeviceDetailView, self).get_context_data(**kwargs)
        devDet = models.Device.objects.filter(pk=self.kwargs.get("pk"))
        context["deviceDetails"] = serializers.serialize("json", devDet)
        return context


@login_required(login_url="/accounts/login/")
def update_device(request):
    print("HI")
    if request.method == "POST":
        instance_form = forms.DeviceUpdateForm(request.POST)
        if instance_form.is_valid():
            instance_form.save()
            return HttpResponseRedirect(reverse_lazy("device_list"))
    else:
        form = forms.DeviceUpdateForm()
    return render(request, "device_update_form.html", {"form": form})


class DeviceUpdateView(LoginRequiredMixin, generic.UpdateView):
    template_name = "device_update_form.html"
    model = models.Device
    form_class = forms.DeviceUpdateForm


@login_required(login_url="/accounts/login/")
def delete_device(request, id):
    context = {}

    obj = get_object_or_404(models.Device, id=id)

    if request.method == "POST":
        obj.status = models.Device.status_bol = 1
        # obj.status = models.Device.status_bol = False
        obj.save()
        return HttpResponseRedirect("/")
    return render(request, "device_delete.html", context)


@login_required(login_url="/accounts/login/")
def create_place(request):
    if request.method == "POST":
        instance_form = forms.PlaceForm(request.POST)
        if instance_form.is_valid():
            instance_form.save()
            return HttpResponseRedirect(reverse_lazy("place_list"))
    else:
        form = forms.PlaceForm()
        return render(request, "place_form.html", {"form": form})


class PlaceList(LoginRequiredMixin, FilterView):
    login_url = "accounts/login"
    paginate_by = 10
    model = models.Place
    context_object_name = "places"
    template_name = "place_list.html"
    filterset_class = filters.PlaceFilter


class PlaceDetailView(LoginRequiredMixin, generic.DetailView):
    login_url = "accounts/login"
    model = models.Place
    template_name = "place_detail.html"
    context_object_name = "place"


class PlaceUpdateView(LoginRequiredMixin, generic.UpdateView):
    template_name = "place_update_form.html"
    model = models.Place
    form_class = forms.PlaceUpdateForm


class PlaceDeleteView(LoginRequiredMixin, DeleteView):
    # login_url = 'accounts/login'
    model = models.Place
    template_name = "place_delete.html"
    success_url = reverse_lazy("place_list")


@login_required(login_url="/accounts/login/")
def create_user_device(request):
    if request.method == "POST":
        instance_form = forms.DeviceUserForm(request.POST)
        if instance_form.is_valid():
            instance_form.save()
            return HttpResponseRedirect(reverse_lazy("user_list"))
    else:
        form = forms.DeviceUserForm()
        return render(request, "user_form.html", {"form": form})


class UserCreateView(LoginRequiredMixin, generic.CreateView):
    # login_url = 'accounts/login'
    model = models.DeviceUser
    template_name = "user_form.html"
    fields = "__all__"

    success_url = reverse_lazy("user_list")


class UserList(LoginRequiredMixin, FilterView):
    # login_url = 'accounts/login'
    paginate_by = 10
    model = models.DeviceUser
    context_object_name = "users"
    template_name = "user_list.html"
    filterset_class = filters.DeviceUserFilter


class DeviceUserDetailView(LoginRequiredMixin, generic.DetailView):
    # login_url = 'accounts/login'
    model = models.DeviceUser
    template_name = "user_detail.html"
    context_object_name = "deviceuser"


class DeviceUserUpdateView(LoginRequiredMixin, generic.UpdateView):
    template_name = "user_update_form.html"
    model = models.DeviceUser
    form_class = forms.DeviceUserUpdateForm


class UserDeleteView(LoginRequiredMixin, generic.DeleteView):
    # login_url = 'accounts/login'
    model = models.DeviceUser
    template_name = "user_delete.html"
    success_url = reverse_lazy("user_list")


@login_required(login_url="/accounts/login/")
def file_upload(request):
    if request.method == "POST":
        form = forms.UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            if request.FILES["file"]:
                utils.handle_uploaded_file(request.FILES["file"])
                return HttpResponseRedirect(reverse_lazy("device_list"))
            else:
                messages.error(request,"Wrong file type!")
                form = forms.UploadFileForm()
                return render(request, "file_upload.html", {"form": form})
        else:
            form = forms.UploadFileForm()
            return render(request, "file_upload.html", {"form": form})
    else:
        form = forms.UploadFileForm()
        return render(request, "file_upload.html", {"form": form})


# def export_devices(request):
#    response = HttpResponse(content_type='application/ms-excel')
#    response['Content-Disposition'] = 'attachment; filename="devices.xls"'

#    wb = xlwt.Workbook(encoding='utf-8')
#    ws = wb.add_sheet('Device Data')  # this will make a sheet named Users Data

# Sheet header, first row
#    row_num = 0

#   font_style = xlwt.XFStyle()
#    font_style.font.bold = True

#    columns = ['Id', 'User History', 'Status', 'Serial number', 'Contract', 'Expiration Date', 'Renewal date',
#               'Host name', 'Make', 'Model', 'Place', 'User']

#    for col_num in range(len(columns)):
#        ws.write(row_num, col_num, columns[col_num], font_style)  # at 0 row 0 column

# Sheet body, remaining rows
#    font_style = xlwt.XFStyle()

#    rows = models.Device.objects.all().values_list('id', 'user_history', 'status', 'serial_number', 'contract',
#                                                   'expiration_date', 'renewal_date', 'host_name', 'make', 'model',
#                                                   'place', 'user')
#    for row in rows:
#        row_num += 1
#        for col_num in range(len(row)):
#            ws.write(row_num, col_num, row[col_num], font_style)

#    wb.save(response)


#    return response
@login_required(login_url="/accounts/login/")
def export_devices(request):
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="devices.xlsx"'

    wb = Workbook()
    ws = wb.active  # this will make a sheet named Users Data

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        "Id",
        "User History",
        "Status",
        "Serial number",
        "Contract",
        "Expiration Date",
        "Renewal date",
        "Host name",
        "Make",
        "Model",
        "Place",
        "User",
    ]

    # for col_num in range(len(columns)):
    #    ws.write(row_num, col_num, columns[col_num], font_style)  # at 0 row 0 column

    # Sheet body, remaining rows
    # font_style = xlwt.XFStyle()

    rows = models.Device.objects.all().values_list(
        "id",
        "user_history",
        "status",
        "serial_number",
        "contract",
        "expiration_date",
        "renewal_date",
        "host_name",
        "make",
        "model",
        "place",
        "user",
    )
    data_list = []
    for row in rows:
        hist_us = ""
        device_user_history = None
        if row[1] is not None:
            device_detail = models.Device.objects.get(pk=row[0])
            device_user_history = device_detail.user_history.all()
            for dev_usr in device_user_history:
                hist_us += dev_usr.name
                hist_us += ","
                hist_us += "/n"

            hist_us[::-2]

            print("Hi")

        status = ""
        if row[2] is not None:
            if row[2] == 1:
                status = "Unavailable"
            elif row[2] == 0:
                status = "Available"

        device_place = ""
        if row[10] is not None:
            device_place = models.Place.objects.get(pk=row[10]).name

        device_user = ""
        if row[11] is not None:
            device_user = models.DeviceUser.objects.get(pk=row[11]).name

        device_expiration_date = ""
        if row[5] is not None:
            device_expiration_date = row[5].strftime("%d/%m/%Y")

        device_renewal_date = ""
        if row[6] is not None:
            device_renewal_date = row[6].strftime("%d/%m/%Y")

        data_row = [
            row[0],
            hist_us,
            status,
            row[3],
            row[4],
            device_expiration_date,
            device_renewal_date,
            row[7],
            row[8],
            row[9],
            device_place,
            device_user,
        ]
        # data_list.append([row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11]])
        ws.append(data_row)
        # row_num += 1
        # for col_num in range(len(row)):
        # ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)

    return response


@login_required(login_url="/accounts/login/")
def export_places(request):
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="places.xls"'

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Places Data")  # this will make a sheet named Users Data

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ["Name", "City", "Address", "CAP", "Country", "Plan"]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)  # at 0 row 0 column

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = models.Place.objects.all().values_list(
        "name", "city", "address", "cap", "country", "plan"
    )
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
    ws.auto_filter.ref = "A:F"

    wb.save(response)

    return response


@login_required(login_url="/accounts/login/")
def export_users(request):
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="users.xls"'

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Users Data")  # this will make a sheet named Users Data

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = [
        "Id",
        "User History",
        "Status",
        "Serial number",
        "Contract",
        "Expiration Date",
        "Renewal date",
        "Host name",
        "Make",
        "Model",
        "Place",
        "User",
    ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)  # at 0 row 0 column

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    rows = models.Device.objects.all().values_list(
        "id",
        "user_history",
        "status",
        "serial_number",
        "contract",
        "expiration_date",
        "renewal_date",
        "host_name",
        "make",
        "model",
        "place",
        "user",
    )
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)

    return response
